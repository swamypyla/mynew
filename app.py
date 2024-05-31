import os
import face_recognition
import cv2
import numpy as np
import streamlit as st
import datetime
import openpyxl


# Function to create Excel database
def create_database(filename):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User_Data"
        ws.append(["Name", "Credit Card Number", "Face Encoding", "CVV", "Expiry Date"])
        wb.save(filename)
        return True
    except Exception as e:
        st.error(f"Error creating database: {e}")
        return False

# Function to add user data to Excel database
def add_user_data(filename, name, credit_card_number, face_encoding, cvv, expiry_date):
    try:
        # Convert face encoding array to a string
        face_encoding_str = ','.join(map(str, face_encoding.tolist()))
        
        wb = openpyxl.load_workbook(filename)
        ws = wb["User_Data"]
        ws.append([name, credit_card_number, face_encoding_str, cvv, expiry_date])
        wb.save(filename)
        return True
    except Exception as e:
        st.error(f"Error adding user data: {e}")
        return False

# Function to retrieve user data from Excel database
def get_user_data(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb["User_Data"]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data
    except Exception as e:
        st.error(f"Error retrieving user data: {e}")
        return []

# Function to check if the provided face matches the registered face
def check_face_match(registered_face_encoding, current_face_encoding):
    try:
        # Convert input variables to numpy arrays if they are not already
        registered_face_encoding = np.array(registered_face_encoding)
        current_face_encoding = np.array(current_face_encoding)

        # Compute the Euclidean distance between the face encodings
        distance = np.linalg.norm(registered_face_encoding - current_face_encoding)

        # You may adjust the tolerance value as needed
        tolerance = 0.6

        # Check if the distance is within the tolerance
        return distance <= tolerance
    except Exception as e:
        st.error(f"Error checking face match: {e}")
        return False

# Streamlit UI
def main():
    st.title("Credit Card Information Access System")

    # Define the filename for the user database
    filename = "user_database.xlsx"

    # Check if the user database file exists, create it if it doesn't
    if not os.path.exists(filename):
        if not create_database(filename):
            st.error("Error creating user database.")
            return

    # Sidebar for adding new user
    st.sidebar.title("Add New User")
    new_name = st.sidebar.text_input("Enter Name:")
    new_credit_card_number = st.sidebar.text_input("Enter Credit Card Number:")
    new_cvv = st.sidebar.text_input("Enter CVV:")
    new_expiry_date = st.sidebar.text_input("Enter Expiry Date (MM/YY):")

    registration_complete = False

    # Check if the "Register Face" button is clicked
    if st.sidebar.button("Register Face"):
        # Capture face encoding
        st.write("Please look at the camera to register your face.")
        face_encoding = capture_face_encoding("registration_photos", "registration_photo.jpg")
        if face_encoding is not None:
            # Add user data to the database
            if add_user_data(filename, new_name, new_credit_card_number, face_encoding, new_cvv, new_expiry_date):
                st.sidebar.success("User registered successfully!")
                registration_complete = True

    # Main content for credit card access
    st.subheader("Credit Card Access")
    credit_card_number_key = "credit_card_number_input"
    credit_card_number = st.text_input("Enter Credit Card Number:", key=credit_card_number_key)
    cvv_input = st.text_input("Enter CVV:", key=f"cvv_input_{registration_complete}", type="password" if registration_complete else "default")
    expiry_date_input = st.text_input("Enter Expiry Date (MM/YY):", key="expiry_date_input", type="password" if registration_complete else "default")

    # Check if the "Access Credit Card Information" button is clicked
    if st.button("Access Credit Card Information"):
        # Check if credit card number is registered
        user_data = get_user_data(filename)
        if user_data:
            credit_card_number_input = credit_card_number.strip()
            cvv_input_stripped = cvv_input.strip()
            expiry_date_input_stripped = expiry_date_input.strip()
            
            credit_card_matched = False
            for user in user_data:
                if user[1].strip() == credit_card_number_input and str(user[3]) == cvv_input_stripped and str(user[4]) == expiry_date_input_stripped:
                    credit_card_matched = True
                    # If registered, capture face and compare with registered face
                    st.write("Please look at the camera for face recognition.")
                    current_face_encoding = capture_face_encoding("access_photos", "authentication_photo.jpg")

                    # Find registered face encoding for the provided credit card number
                    registered_face_encoding = np.fromstring(user[2], dtype=float, sep=',')

                    # Check if faces match
                    if check_face_match(registered_face_encoding, current_face_encoding):
                        st.success("Face matched. Access granted.")
                        # Show credit card details
                        st.write(f"Name: {user[0]}")
                        st.write(f"Credit Card Number: {user[1]}")
                        st.write(f"CVV: {user[3]}")
                        st.write(f"Expiry Date: {user[4]}")
                        break
                    else:
                        st.error("Fraud detected! Face doesn't match registered user.")
                        # Capture image of unauthorized user and save in "fraud" folder
                        capture_unauthorized_user()
                    break
            
            if not credit_card_matched:
                st.error("Credit card details not found!")
        else:
            st.error("No users registered yet.")


# Function to capture face encoding
def capture_face_encoding(folder_path, filename):
    try:
        # Capture video frame
        cap = cv2.VideoCapture(0)
        ret, frame = cap.read()

        # Convert to RGB (face_recognition requires RGB format)
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # Find face location and encoding
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        if face_encodings:
            # Save the photo in the specified folder
            photo_path = os.path.join(folder_path, filename)
            cv2.imwrite(photo_path, frame)
            return np.array(face_encodings[0])
        else:
            st.error("No face detected! Please try again.")
            return None
    except Exception as e:
        st.error(f"Error capturing face encoding: {e}")
        return None

# Function to capture image of unauthorized user and save in "fraud" folder
def capture_unauthorized_user():
    try:
        # Capture video frame
        cap = cv2.VideoCapture(0)
        ret, frame = cap.read()

        # Save image in "fraud" folder
        if not os.path.exists("fraud"):
            os.makedirs("fraud")
        image_path = os.path.join("fraud", f"unauthorized_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.png")
        cv2.imwrite(image_path, frame)
    except Exception as e:
        st.error(f"Error capturing unauthorized user: {e}")


if __name__ == "__main__":
    main()

